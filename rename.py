import os
from os import listdir
from os.path import isfile, isdir, join
import json
import pandas as pd
import openpyxl
import shutil
import eyed3	

df_Spec = pd.DataFrame(columns=['Year', 'No', 'Title', 'Artist', 'url', 'htFilename'])
df_Physical = pd.DataFrame(columns=['url', 'Filename'])


def getUrls(strExcelName, strSheetName):
	global df_Spec

	workbook = openpyxl.load_workbook(strExcelName, data_only=True)
	st = workbook[strSheetName]

	print('Sheet : ' + strSheetName)

	for row in st.iter_rows(min_row=0):
		if len(row) > 3 :
			if str(row[3].value).startswith('https'):
				df_Spec = df_Spec.append({'Year':strSheetName, 'No':str(int(row[0].value)).zfill(3), 'Title':str(row[1].value), 'Artist':str(row[2].value), 'url':str(row[3].value)}, ignore_index=True)
			
	
def getPhysical(strHDT_Filename):
	global df_Physical

	with open(strHDT_Filename) as f:
		datas = json.load(f)
	
	for data in datas:
		if 'gal_num' in data and 'names' in data:
			df_Physical = df_Physical.append({'url':data['gal_num'],'Filename':data['names'][0]}, ignore_index=True)
		
	#print(df_Physical)
	
def mapping():
	global df_Spec
	global df_Physical
	
	for index, row in df_Physical.iterrows():
		res = df_Spec['url'][df_Spec['url']==row['url']].index.tolist()
		df_Spec.loc[res, 'htFilename'] = row['Filename']
		#print(res)
		
def rename():
	workbook = openpyxl.load_workbook('final_mapping.xlsx', data_only=True)
	st = workbook['sheet1']
	
	strBaseDir = r'C:\Users\Tony\Downloads\hitomi_downloader_GUI\Billboard100'
	
	for row in st.iter_rows(min_row=2):
		if row[0] != 'x':
			iYear = row[1].value
			iNo = int(row[2].value)
			strNo = str(int(row[2].value)).zfill(3)
			strTitle = (str(row[3].value).replace('?','')).replace(':', ' ')
			strArtist = row[4].value
			strHTFilename = row[6].value
			
			strTargetDir = os.path.join(strBaseDir, 'Billboard100_'+str(iYear))
			if not os.path.exists(strTargetDir):
				os.makedirs(strTargetDir)

			strSourceFilename = os.path.join(r'C:\Users\Tony\Downloads\hitomi_downloader_GUI', strHTFilename)
			strTargetFilename = os.path.join(strTargetDir, strNo+'.'+strTitle+'.mp3')
			
			print(strSourceFilename, strTargetFilename)
			shutil.copyfile(strSourceFilename, strTargetFilename)
			
def change_audio_tag():
	workbook = openpyxl.load_workbook('final_mapping.xlsx', data_only=True)
	st = workbook['sheet1']
	
	strBaseDir = r'C:\Users\Tony\Downloads\hitomi_downloader_GUI\Billboard100'
	
	for row in st.iter_rows(min_row=2):
		if row[0] != 'x':
			iYear = row[1].value
			iNo = int(row[2].value)
			strNo = str(int(row[2].value)).zfill(3)
			strTitle = (str(row[3].value).replace('?','')).replace(':', ' ')
			strArtist = row[4].value
			strHTFilename = row[6].value
			
			strTargetDir = os.path.join(strBaseDir, 'Billboard100_'+str(iYear))
			if not os.path.exists(strTargetDir):
				os.makedirs(strTargetDir)

			strSourceFilename = os.path.join(r'C:\Users\Tony\Downloads\hitomi_downloader_GUI', strHTFilename)
			strTargetFilename = os.path.join(strTargetDir, strNo+'.'+strTitle+'.mp3')
	
			print(strTargetFilename)
			audio = eyed3.load(strTargetFilename)
	
			audio.tag.artist = strArtist
			audio.tag.album = 'Billboard Top 100'
			audio.tag.title = strTitle
			audio.tag.track_num = (iNo, None)
	
			audio.tag.save()
	

if __name__ == "__main__":

# prepare physical
	getPhysical('billboard.hdt')
	#getPhysical('test.hdt')

# prepare spec or read spec
	'''
	for iYear in range(1960, 2003):
	#for iYear in range(1970, 1972):
		getUrls('Billboards Top 100.xlsx', str(iYear))
		
	df_Spec.to_csv('df_Spec.csv', sep ='\t')
	'''
	'''
	df_Spec = pd.read_csv('df_Spec.csv', sep='\t')
	#print(df_Spec)
	
	mapping()
	print(df_Spec)
	df_Spec.to_csv('mapping.csv', sep ='\t')
	'''
	
	#rename()
	
	change_audio_tag()
	





	
